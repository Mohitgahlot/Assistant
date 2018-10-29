########transformation applied on -> (process applied) -> variable that has the final output of the process########

#from original sentence -> (stemming) -> stemmed
#from original sentence -> (lemmatization) -> lemmatized
#from original sentence -> (stop word removal) -> filtered_sentence

########FUNCTIONS########
#lookup_words(string) -> string with acronyms/slangs removed        to be used only after stopwords have been removed and punctuations as well

#STOPWORD REMOVAL and REMOVING PUNCTUATIONS
import nltk
import sys
#nltk.download('wordnet')
#nltk.download('averaged_perceptron_tagger')
#nltk.download('stopwords')
from nltk.corpus import stopwords
#from nltk.tokenize import word_tokenize

example_sent = sys.argv[1]
def basic_processing(string):
        stop_words = set(stopwords.words('english')) - {'where'}
        #word_tokens = word_tokenize(example_sent)
        import re
        word_tokens= re.split('; |, |\*|\n|\.|\?|\s',string.lower())
        #filtered_sentence = [w for w in word_tokens if not w in stop_words]
        filtered_sentence = []
        for w in word_tokens:
                if w not in stop_words and w is not '':
                        filtered_sentence.append(w)
        filtered_sentence = ' '.join(filtered_sentence)
        return filtered_sentence

#REMOVAL OF ACRONYMS AND SLANGS

def lookup_words(input_text):
        global lookup_dict

        words = input_text.split()
        new_words = []

        for word in words:
                if word.lower() in lookup_dict:
                    word = lookup_dict[word.lower()]
                new_words.append(word)

        new_text = " ".join(new_words)
        return new_text

# INPUT -> lookup_words("Does the cllg have a canteen")
# OUTPUT ->'Does the college have a canteen'

#AUTOCORRECT WORDS

#from autocorrect import spell

def autocorrect(string):
	sentence = []
	for word in string.split():
		word = spell(word)
		sentence.append(word)

	correct_sentence = ' '.join(sentence)
	return correct_sentence

# INPUT -> autocorrect("Thenks alexa")
# OUTPUT ->'Thanks Alexa'

#STEMMING & LEMMATIZATION

def stem_lemm(string):
	global stemmed
	global lemmatized
	porter_stemmer = PorterStemmer()
	lem = WordNetLemmatizer()

	for word in string.split():
		w_stem = porter_stemmer.stem(word)
		stemmed += w_stem + ' '

		w_lem = lem.lemmatize(word, "v")
		lemmatized += w_lem + ' '
	return lemmatized

#what is being talked about?
from nltk import pos_tag
def pos_tagging(string):
	tokens = string.split()
	POS = pos_tag(tokens)

	subject = []

	for k,v in POS:
		if v.find('NN') != - 1:
			subject.append(k)

	return subject

#pos tagging apply before or after stop word

lw = 'NONE'
ac = 'NONE'
st = 'NONE'
bp = basic_processing(example_sent)

type = sys.argv[2]

lookup_dict = {'der':'there', 'btw':'by the way', "awsm" : "awesome", "cllg" :"college", "abt": "about" , 'gr8':'great', 'np':'no problem', 'okiee':'okay', 'omg':'oh my god' , '2day':'today'}
from nltk.stem.porter import PorterStemmer
from nltk.stem.wordnet import WordNetLemmatizer

stemmed = ''
lemmatized = ''

if(type==1):
        lw = lookup_words(bp)
        ac = autocorrect(lw)
        st = stem_lemm(ac)
else:
        st = stem_lemm(bp)

subject = pos_tagging(st)
#print(subject)
#print(st)

import openpyxl

#subject = ['qualification', 'faculty', 'members']
#st = 'qualification faculty members ncu'
st = st.split()
wb = openpyxl.load_workbook('/Applications/XAMPP/xamppfiles/htdocs/Assistant/DB_1.xlsx')
ques_mapping = wb['Sheet1']
f = 0
ans_id = -1
for r in range(2,39):

    concept = ques_mapping.cell(row=r, column=2).value
    concept_id = ques_mapping.cell(row=r, column=1).value
    feature = ques_mapping.cell(row=r, column=4).value

    feature = feature.split(';')

    if concept in subject:
        ans_id = 0
        for f in feature:
            if f in subject or f in st:
                ans_id = ques_mapping.cell(row=r, column=5).value
                f = 1
                break;
        if f==1:
            break;

    if ans_id<=0:
        for f in feature:
            if f in subject or f in st:
                ans_id = ques_mapping.cell(row=r, column=5).value
                f = 1
                break;

#print(ans_id)
answer = "NA"
redirect = "NA"
display = "NA"
if ans_id == 0:
    answer = "I'm afraid I do not have an answer to that question!"

elif ans_id == -1:
    answer = "I'm afraid I did not understand your question!"
else:
    wb = openpyxl.load_workbook('/Applications/XAMPP/xamppfiles/htdocs/Assistant/DB_4.xlsx')
    ans_mapping = wb['Sheet1']
    for  r in range(2,29):
        id = ans_mapping.cell(row=r, column=1).value
        if id == ans_id:
            answer = ans_mapping.cell(row=r, column=2).value
            redirect = ans_mapping.cell(row=r, column=3).value
            display = ans_mapping.cell(row=r, column=4).value
            break;
print(answer+"**"+redirect)
#print(answer)
#file = open("/Applications/XAMPP/xamppfiles/htdocs/Assistant/result.txt","w")
#file.write(answer)
#file.write("\n")
#file.write(redirect)
#file.write("\n")
#file.write(display)
#file.close()




